from selenium import webdriver
from time import sleep
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

import pandas as pd
from os import remove
import re
from openpyxl import load_workbook
from openpyxl.styles import Font

# time at start of program
start_time = time.time()

options = webdriver.ChromeOptions() 
#options.add_argument("user-data-dir=C:\\Users\\mike\\AppData\\Local\\Google\\Chrome\\User Data") #Path to your chrome profile
#options.add_argument('--profile-directory=Profile 2')

# path to chrome driver exe
driver = webdriver.Chrome(ChromeDriverManager().install())

# implicit wait
driver.implicitly_wait(10)

# Open url file to be read
top_keyword = ''
file = open('./urls.txt', 'r')

# function takes a text file and returns a dictionary of URLs
def text2dict(file_obj):
    """ Takes file object as input and return python dictionary object"""
    keys = []
    values = []
    contents = file_obj.read()
    contents_as_row = contents.split('\n')

    for line in contents_as_row:
        if not re.match("^http.*://", line):
            keys.append(line)
            values.append([])
        else:
            values[-1].append(line)

    new_dict = dict(list(zip(keys, values)))

    return new_dict

# Go to chrome downloads page
def every_downloads_chrome(driver):
    if not driver.current_url.startswith("chrome://downloads"):
        driver.get("chrome://downloads/")
    return driver.execute_script("""
        return document.querySelector('downloads-manager')
        .shadowRoot.querySelector('#downloadsList')
        .items.filter(e => e.state === 'COMPLETE')
        .map(e => e.filePath || e.file_path || e.fileUrl || e.file_url);
        """)

# Download the CSV files from the URLs given
def main(url):
    # Go to site
    driver.get("https://ahrefs.com/user/login")
    sleep(2)

    # Check if already logged in or not
    if "Sign in to Ahrefs" in driver.page_source:                        #Login To Ahrefs
        # Enter credentials to sign in
        driver.find_element_by_name("email").send_keys("jasmin-i@windowslive.com")                #Insert Email
        sleep(2)
        driver.find_element_by_name("password").send_keys("jasmin020894")             #Insert Password
        sleep(2)
        driver.find_element_by_xpath('//*[@id="root"]/div/div/div[1]/div/div/div/form/div/button/div').click()

    # Insert the URL into the search/input box
    driver.find_element_by_class_name("css-j3q2yv-input").send_keys(url)
    sleep(1)
    # Click enter
    driver.find_element_by_class_name("css-184i2sh-button").click()

    # Select pre organic filter
    try:
        element = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, "pe_organic_keywords")))
        element.click()
    except:
        driver.find_element_by_class_name("css-184i2sh-button").click()
        driver.find_element_by_id("pe_organic_keywords").click()

    sleep(3)

    # Return if no results
    if "There are no results in our index for the specified domain/URL." in driver.page_source:
        return
    if "No organic keywords were found." in driver.page_source:
        return

    # Filter the position title
    #try:
    driver.find_element_by_id("filter-position-title").click()
    #except:
    #    driver.find_element_by_id("pe_organic_keywords").click()
    #    driver.find_element_by_id("filter-position-title").click()

    sleep(2)
    # Enter 10 into filter position
    driver.find_element_by_id("filter-position-to").send_keys("10")
    sleep(1)
    # Click enter
    if "No organic keywords were found." in driver.page_source:
        return
    
    driver.find_element_by_css_selector("#drop_down_menu_position > .dropdown-menu .btn").click()

    # Export the CSV
    try:
        element = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable((By.ID, "export_button")))
        element.click()
    except:
        return

    sleep(2)
    try:
       driver.find_element_by_css_selector("#btn_export_type_utf-16").click()
    except:
        return

    sleep(2)

    # Click start export button
    try:
        element = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "start_export_button")))
        element.click()
    except:
        driver.refresh()
        element = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable((By.ID, "export_button")))
        element.click()
        sleep(1)
        driver.find_element_by_css_selector("#btn_export_type_utf-16").click()
        sleep(1)
        driver.find_element_by_id("start_export_button").click()


    # Wait and return path to export
    sleep(2)
    path = (WebDriverWait(driver, 120, 1).until(every_downloads_chrome))[0]
    return path

# Loop through list of URLs and store paths of CSVs in a list
def download_list(key,urls):
    global top_keyword
    top_keyword = str(key)+'.csv'
    downloads = []
    # loop URLs and download CSVs
    for url in urls:
        try:
            downloads.append(main(url))
        except:
            downloads.append(main(url))
    # return non-empty list of path to downloads
    return [down for down in downloads if down is not None]

# Merge all downloaded CSV's into one CSV
def merge_csvs(merge_list, file_name):    #merges all csvs for each main keyword into one big csv
    if len(merge_list) > 0:
        # combine all files in the list
        combined_csv = pd.concat([pd.read_csv(f, encoding= 'utf-16', sep= '\t') for f in merge_list])
        # export to csv
        combined_csv.to_csv(file_name, index=False)

        # deletes csvs that were used for merging

        for fname in merge_list:   
            remove(fname)
        # Return merged file
        return file_name
    
# Function to clean up merged file
def remove_columns(file):

    try:
        f = pd.read_csv(file)
        # Filter for these columns
        keep_col = ['Keyword', 'Volume', 'Page URL inside']
        new_f = f[keep_col]

        new_f = new_f[new_f.Keyword != 'Keyword'] #Removes duplicate headers
        new_f = new_f.sort_values('Page URL inside')
        new_f = new_f.drop_duplicates(subset='Keyword')
        new_f.Volume = pd.to_numeric(new_f.Volume, errors='coerce') #converts Volume column to numbers
        new_f = new_f.sort_values('Volume', ascending=[False])

        new_f = new_f.reset_index(drop=True)

        d = {'x': ['', '', '', '', '', '', '', ''], 'URL': ['Page Title (H1)', 'Choco Lite - in Pareri Forum si Farmacii, Pret Chocolite (2020)', 'SEO Title Tag', 'Choco Lite - in Pareri Forum si Farmacii, Pret (2020)', 'Permalink', 'choco-lite-pareri-forum', 'Meta Description', 'Cititi Cele Mai Noi Informatii Despre Suplimentul Alimentar Choco Lite in %currentyear%. Aflati Cum Transforma Procesul Dificil De…']}
        df = pd.DataFrame(data=d, index=[0,1,2,3,4,5,6, 7])
        new_ff = pd.concat([new_f, df], axis=1, join='outer')

        # Save to csv
        new_ff.to_csv(file, index=False)
        # read csv into pandas again
        read_file = pd.read_csv(file)

        # get max length of each column
        col1_len = read_file['Keyword'].astype(str).str.len().max()
        col2_len = read_file['Volume'].astype(str).str.len().max()
        col3_len = read_file['Page URL inside'].astype(str).str.len().max()
        col4_len = read_file['x'].astype(str).str.len().max()
        col5_len = read_file['URL'].astype(str).str.len().max()

        # remove csv
        file_name = file.split('.')[0]

        # save output to excel
        read_file.to_excel(file_name + '.xlsx', index=None, header=True)


        wb = load_workbook(file_name + '.xlsx')

        ws = wb['Sheet1']

        # Set column widths
        ws.column_dimensions['A'].width = col1_len
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = col3_len
        ws.column_dimensions['D'].width = col4_len
        ws.column_dimensions['E'].width = col5_len

        ws['E2'].font = Font(bold=True)
        ws['E4'].font = Font(bold=True)
        ws['E6'].font = Font(bold=True)
        ws['E8'].font = Font(bold=True)

        wb.save(file_name + '.xlsx')
        
        # remove the .csv duplicate file.
        old_file = str(file_name) + '.csv'
        print("old_file is : ", old_file)
        remove(old_file)

    except:
        return

def autofit_cols(filename, dfs):
    options = {}
    options['strings_to_formulas'] = False
    options['strings_to_urls'] = False
    writer = pd.ExcelWriter(filename, engine='xlsxwriter',options=options)
    for sheetname, df in dfs.items():  # loop through `dict` of dataframes
        df.to_excel(writer, sheet_name=sheetname, index=False)  # send df to writer
        worksheet = writer.sheets[sheetname]  # pull worksheet object
        for idx, col in enumerate(df):  # loop through all columns
            series = df[col]
            max_len = max((
                series.astype(str).map(len).max(),  # len of largest item
                len(str(series.name))  # len of column name/header
            )) + 1  # adding a little extra space
            worksheet.set_column(idx, idx, max_len)  # set column width
    writer.save()

# Get a dictionary of URLs from the text file
my_dict = text2dict(file)

print("my_dict is : \n", my_dict)

# Loop through each dictionary key and get merged exports from URL search
for x in my_dict:
    print (x, end = ':')
    remove_columns(merge_csvs(download_list(x, my_dict[x]), top_keyword))
    print ('Complete')

driver.close()

# Print how long the program took
print("Program took --- %s seconds --- to run!" % (time.time() - start_time))
